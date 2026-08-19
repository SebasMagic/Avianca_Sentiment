"""
Excel writer — exporta las menciones a un .xlsx con formato.
Archivo generado: avianca_mentions_YYYY-MM-DD.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timezone
from config import BRAND_KEYWORD


COLUMNS = [
    ("¿Queja real?",      "is_complaint"),
    ("Plataforma",        "platform"),
    ("Texto",             "text"),
    ("Autor",             "author"),
    ("Fecha publicación", "published_at"),
    ("URL",               "source_url"),
    ("Likes",             "likes"),
    ("Shares",            "shares"),
    ("Comentarios",       "comments_count"),
    ("Sentiment +",       "sentiment_positive"),
    ("Sentiment -",       "sentiment_negative"),
    ("Sentiment =",       "sentiment_neutral"),
    ("Emoción",           "emotion"),
    ("Extraído en",       "fetched_at"),
]

# Colores por plataforma
PLATFORM_COLORS = {
    "web":       "D6E4F0",
    "instagram": "FCE4EC",
    "tiktok":    "E0F7FA",
    "twitter":   "E3F2FD",
}

HEADER_FILL  = PatternFill("solid", fgColor="0D1117")
HEADER_FONT  = Font(name="Calibri", bold=True, color="F97316", size=10)


def export(mentions: list[dict]) -> str:
    """
    Escribe las menciones a un Excel y retorna la ruta del archivo generado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menciones"

    # ── Encabezado
    headers = [col[0] for col in COLUMNS]
    ws.append(headers)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    COMPLAINT_FILL   = PatternFill("solid", fgColor="FFE5E5")   # rojo suave
    COMPLAINT_FONT   = Font(name="Calibri", bold=True, color="C0392B", size=10)
    NO_COMPLAINT_FONT = Font(name="Calibri", color="27AE60", size=10)

    # ── Filas
    for mention in mentions:
        is_complaint = mention.get("is_complaint", False)
        row = []
        for _, field in COLUMNS:
            val = mention.get(field, "")
            # Convertir booleano a texto legible
            if field == "is_complaint":
                val = "SÍ ⚠️" if val else "no"
            row.append(val)
        ws.append(row)

        row_idx = ws.max_row
        platform = mention.get("platform", "")

        # Quejas reales: fondo rojo suave en toda la fila
        if is_complaint:
            fill = COMPLAINT_FILL
        else:
            bg_color = PLATFORM_COLORS.get(platform, "FFFFFF")
            fill = PatternFill("solid", fgColor=bg_color)

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            # Columna is_complaint: fuente con color
            if col_idx == 1:
                cell.font = COMPLAINT_FONT if is_complaint else NO_COMPLAINT_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 3))
            cell.border = Border(bottom=Side(style="thin", color="E0E0E0"))

    # ── Anchos de columna
    col_widths = [13, 14, 60, 22, 22, 45, 9, 9, 14, 13, 13, 13, 14, 22]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # ── Freeze header
    ws.freeze_panes = "A2"

    # ── Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Nombre del archivo con fecha
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"{BRAND_KEYWORD.lower()}_mentions_{date_str}.xlsx"
    wb.save(filename)

    print(f"[Excel] Archivo guardado: {filename}  ({len(mentions)} filas)")
    return filename
